"""
Web Scraper Profesional
=======================

Un sistema de web scraping versátil y robusto que incluye:
- Extracción de datos de páginas web
- Scraping de noticias
- Monitoreo de precios de productos
- Exportación a múltiples formatos (JSON, CSV, Excel)
- Sistema de programación de tareas

Autor: [Tu Nombre]
Fecha: 2026
"""

# =============================================================================
# IMPORTACIONES
# =============================================================================

import json
import csv
import os
import re
import time
import logging
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict
from urllib.parse import urljoin, urlparse

# Librerías para scraping
import httpx
import asyncio
from bs4 import BeautifulSoup

# Para exportación a Excel (opcional)
try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

# =============================================================================
# CONFIGURACIÓN DE LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURACIÓN GLOBAL
# =============================================================================

# Headers para simular un navegador real (evitar bloqueos)
HEADERS_NAVEGADOR = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'es-ES,es;q=0.8,en-US;q=0.5,en;q=0.3',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
}

# Tiempo de espera entre peticiones (para ser respetuoso)
TIEMPO_ESPERA_SEGUNDOS = 2

# Directorio para guardar resultados
DIRECTORIO_SALIDA = "scraping_resultados"

# =============================================================================
# CLASES DE DATOS
# =============================================================================

@dataclass
class ResultadoScraping:
    """Estructura para almacenar un resultado de scraping"""
    url: str
    titulo: str
    contenido: Dict[str, Any]
    fecha_extraccion: str
    fuente: str
    metadata: Optional[Dict] = None


@dataclass
class ProductoEcommerce:
    """Estructura para productos de e-commerce"""
    nombre: str
    precio: str
    precio_numerico: Optional[float]
    moneda: str
    disponibilidad: str
    url: str
    imagen_url: Optional[str]
    tienda: str
    fecha_consulta: str

# =============================================================================
# FUNCIONES DE UTILIDAD
# =============================================================================

def crear_directorio_salida() -> None:
    """Crea el directorio de salida si no existe"""
    if not os.path.exists(DIRECTORIO_SALIDA):
        os.makedirs(DIRECTORIO_SALIDA)
        logger.info(f"Directorio creado: {DIRECTORIO_SALIDA}")


def limpiar_texto(texto: str) -> str:
    """
    Limpia el texto eliminando espacios extra y caracteres no deseados.
    """
    if not texto:
        return ""
    # Eliminar espacios múltiples y saltos de línea
    texto = re.sub(r'\s+', ' ', texto)
    # Eliminar espacios al inicio y final
    return texto.strip()


def extraer_precio_numerico(texto_precio: str) -> Optional[float]:
    """
    Extrae el valor numérico de un texto de precio.
    Soporta múltiples formatos: $1,234.56, 1.234,56€, etc.
    """
    if not texto_precio:
        return None
    
    # Eliminar símbolos de moneda y espacios
    texto_limpio = re.sub(r'[^\d.,]', '', texto_precio)
    
    # Detectar formato (europeo vs americano)
    if ',' in texto_limpio and '.' in texto_limpio:
        # Formato: 1,234.56 o 1.234,56
        if texto_limpio.rfind(',') > texto_limpio.rfind('.'):
            # Formato europeo: 1.234,56
            texto_limpio = texto_limpio.replace('.', '').replace(',', '.')
        else:
            # Formato americano: 1,234.56
            texto_limpio = texto_limpio.replace(',', '')
    elif ',' in texto_limpio:
        # Podría ser 1,234 (miles) o 1,23 (decimal)
        if len(texto_limpio.split(',')[1]) <= 2:
            # Probablemente decimal europeo
            texto_limpio = texto_limpio.replace(',', '.')
        else:
            # Probablemente separador de miles americano
            texto_limpio = texto_limpio.replace(',', '')
    
    try:
        return float(texto_limpio)
    except ValueError:
        return None


def obtener_dominio(url: str) -> str:
    """Extrae el dominio de una URL"""
    parsed = urlparse(url)
    return parsed.netloc

# =============================================================================
# CLASE PRINCIPAL: SCRAPER BASE
# =============================================================================

class ScraperBase:
    """
    Clase base para todos los scrapers asíncronos.
    """
    
    def __init__(self, headers: Optional[Dict] = None):
        self.headers = headers or HEADERS_NAVEGADOR
        self.client = httpx.AsyncClient(headers=self.headers, timeout=30.0, follow_redirects=True)
        self.config = self._cargar_configuracion()

    def _cargar_configuracion(self) -> Dict:
        """Carga la configuración desde el archivo JSON externo"""
        config_path = os.path.join(os.path.dirname(__file__), "scraper_config.json")
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error cargando configuración: {e}")
            return {}

    async def cerrar(self):
        """Cierra el cliente HTTP"""
        await self.client.aclose()
        
    async def obtener_pagina(self, url: str, reintentos: int = 3) -> Optional[BeautifulSoup]:
        """Obtiene el contenido de una página web de forma asíncrona"""
        for intento in range(reintentos):
            try:
                logger.info(f"Obteniendo: {url} (intento {intento + 1}/{reintentos})")
                response = await self.client.get(url)
                response.raise_for_status()
                
                return BeautifulSoup(response.content, 'html.parser')
                
            except httpx.HTTPError as e:
                logger.error(f"Error obteniendo {url}: {e}")
                if intento < reintentos - 1:
                    await asyncio.sleep(2 * (intento + 1))
                    
        return None
    
    def extraer_texto_elemento(self, soup: BeautifulSoup, selector: str, 
                                atributo: Optional[str] = None) -> Optional[str]:
        """
        Extrae texto o atributo de un elemento usando selector CSS.
        
        Args:
            soup: Objeto BeautifulSoup
            selector: Selector CSS (ej: 'h1.title', '.price', '#content')
            atributo: Atributo a extraer (ej: 'href', 'src'). Si es None, extrae texto.
            
        Returns:
            Texto o valor del atributo, o None si no se encuentra
        """
        elemento = soup.select_one(selector)
        if elemento:
            if atributo:
                return elemento.get(atributo)
            return limpiar_texto(elemento.get_text())
        return None
    
    def extraer_lista_elementos(self, soup: BeautifulSoup, selector: str) -> List[str]:
        """
        Extrae una lista de textos de múltiples elementos.
        
        Args:
            soup: Objeto BeautifulSoup
            selector: Selector CSS para múltiples elementos
            
        Returns:
            Lista de textos extraídos
        """
        elementos = soup.select(selector)
        return [limpiar_texto(elem.get_text()) for elem in elementos if elem.get_text()]

# =============================================================================
# SCRAPER DE NOTICIAS
# =============================================================================

class ScraperNoticias(ScraperBase):
    """
    Scraper especializado para extraer noticias de sitios web.
    Soporta múltiples fuentes y estructura de datos estandarizada.
    """
    
    def obtener_selectores(self, url: str) -> Dict:
        """Obtiene los selectores desde la configuración externa"""
        noticias_config = self.config.get('noticias', {})
        dominio = obtener_dominio(url)
        for sitio, selectores in noticias_config.items():
            if sitio in dominio:
                return selectores
        return noticias_config.get('default', {})
    
    async def extraer_noticias(self, url: str, max_noticias: int = 10) -> List[Dict]:
        soup = await self.obtener_pagina(url)
        if not soup:
            return []
        
        selectores = self.obtener_selectores(url)
        noticias = []
        
        # Buscar contenedores de artículos
        articulos = soup.select(selectores['articulos'])[:max_noticias]
        
        for articulo in articulos:
            try:
                noticia = self._extraer_datos_articulo(articulo, selectores, url)
                if noticia and noticia.get('titulo'):
                    noticias.append(noticia)
            except Exception as e:
                logger.error(f"Error extrayendo artículo: {e}")
                continue
        
        logger.info(f"Extraídas {len(noticias)} noticias de {url}")
        return noticias
    
    def _extraer_datos_articulo(self, articulo: BeautifulSoup, 
                                 selectores: Dict, url_base: str) -> Dict:
        """Extrae los datos de un artículo individual"""
        titulo_elem = articulo.select_one(selectores['titulo'])
        resumen_elem = articulo.select_one(selectores['resumen'])
        enlace_elem = articulo.select_one(selectores['enlace'])
        fecha_elem = articulo.select_one(selectores['fecha'])
        imagen_elem = articulo.select_one(selectores['imagen'])
        
        # Construir URL completa si es relativa
        enlace = None
        if enlace_elem and enlace_elem.get('href'):
            enlace = urljoin(url_base, enlace_elem.get('href'))
        
        return {
            'titulo': limpiar_texto(titulo_elem.get_text()) if titulo_elem else None,
            'resumen': limpiar_texto(resumen_elem.get_text()) if resumen_elem else None,
            'enlace': enlace,
            'fecha': fecha_elem.get('datetime') or limpiar_texto(fecha_elem.get_text()) if fecha_elem else None,
            'imagen': imagen_elem.get('src') if imagen_elem else None,
            'fuente': obtener_dominio(url_base),
            'fecha_extraccion': datetime.utcnow().isoformat()
        }
    
    async def extraer_articulo_completo(self, url: str) -> Optional[Dict]:
        """
        Extrae el contenido completo de un artículo.
        
        Args:
            url: URL del artículo
            
        Returns:
            Diccionario con el contenido completo del artículo
        """
        soup = await self.obtener_pagina(url)
        if not soup:
            return None
        
        # Extraer título
        titulo = self.extraer_texto_elemento(soup, 'h1')
        
        # Extraer contenido del artículo
        parrafos = soup.select('article p, .article-content p, .content p, [class*="article"] p')
        contenido = '\n\n'.join([limpiar_texto(p.get_text()) for p in parrafos if p.get_text()])
        
        # Extraer autor
        autor = self.extraer_texto_elemento(soup, '[class*="author"], .author, .byline')
        
        # Extraer fecha
        fecha = self.extraer_texto_elemento(soup, 'time, [class*="date"]')
        
        return {
            'url': url,
            'titulo': titulo,
            'autor': autor,
            'fecha': fecha,
            'contenido': contenido,
            'fuente': obtener_dominio(url),
            'fecha_extraccion': datetime.utcnow().isoformat()
        }

# =============================================================================
# SCRAPER DE E-COMMERCE (PRECIOS)
# =============================================================================

class ScraperEcommerce(ScraperBase):
    """
    Scraper especializado para monitorear precios de productos.
    Soporta múltiples tiendas online.
    """
    
    def identificar_tienda(self, url: str) -> str:
        """Identifica la tienda basándose en la URL"""
        dominio = obtener_dominio(url).lower()
        if 'amazon' in dominio:
            return 'amazon'
        elif 'mercadolibre' in dominio:
            return 'mercadolibre'
        return 'default'
    
    def obtener_patrones(self, url: str) -> Dict:
        """Obtiene los patrones desde la configuración externa"""
        tienda = self.identificar_tienda(url)
        ecommerce_config = self.config.get('ecommerce', {})
        return ecommerce_config.get(tienda, ecommerce_config.get('default', {}))

    async def extraer_producto(self, url: str) -> Optional[ProductoEcommerce]:
        soup = await self.obtener_pagina(url)
        if not soup:
            return None
        
        patrones = self.obtener_patrones(url)
        
        # Extraer nombre
        nombre = self.extraer_texto_elemento(soup, patrones['nombre'])
        
        # Extraer precio
        precio_texto = self.extraer_texto_elemento(soup, patrones['precio'])
        precio_numerico = extraer_precio_numerico(precio_texto)
        
        # Detectar moneda
        moneda = 'USD'
        if precio_texto:
            if '€' in precio_texto:
                moneda = 'EUR'
            elif 'S/' in precio_texto or 'S/. ' in precio_texto:
                moneda = 'PEN'
            elif '$' in precio_texto:
                moneda = 'USD'
            elif 'R$' in precio_texto:
                moneda = 'BRL'
        
        # Extraer disponibilidad
        disponibilidad = self.extraer_texto_elemento(soup, patrones['disponibilidad']) or 'No disponible'
        
        # Extraer imagen
        imagen_url = self.extraer_texto_elemento(soup, patrones['imagen'], 'src')
        
        return ProductoEcommerce(
            nombre=nombre or 'Producto sin nombre',
            precio=precio_texto or 'Precio no disponible',
            precio_numerico=precio_numerico,
            moneda=moneda,
            disponibilidad=disponibilidad,
            url=url,
            imagen_url=imagen_url,
            tienda=obtener_dominio(url),
            fecha_consulta=datetime.utcnow().isoformat()
        )
    
    async def monitorear_productos(self, urls: List[str]) -> List[ProductoEcommerce]:
        """
        Monitorea múltiples productos de forma concurrente.
        """
        tareas = [self.extraer_producto(url) for url in urls]
        resultados = await asyncio.gather(*tareas)
        return [r for r in resultados if r]

# =============================================================================
# SCRAPER GENÉRICO
# =============================================================================

class ScraperGenerico(ScraperBase):
    """
    Scraper genérico para extraer datos de cualquier página web.
    Útil para casos específicos donde se conocen los selectores.
    """
    
    async def extraer_datos(self, url: str, configuracion: Dict) -> Dict:
        """Extrae datos según una configuración de selectores (Async)"""
        soup = await self.obtener_pagina(url)
        if not soup:
            return {'error': 'No se pudo obtener la página'}
        
        resultado = {
            'url': url,
            'fecha_extraccion': datetime.utcnow().isoformat()
        }
        
        for campo, selector in configuracion.items():
            if isinstance(selector, dict):
                # Selector con atributo
                valor = self.extraer_texto_elemento(
                    soup, 
                    selector.get('selector'), 
                    selector.get('atributo')
                )
            else:
                # Selector simple de texto
                valor = self.extraer_texto_elemento(soup, selector)
            
            resultado[campo] = valor
        
        return resultado
    
    async def extraer_tabla(self, url: str, selector_tabla: str) -> List[Dict]:
        soup = await self.obtener_pagina(url)
        if not soup:
            return []
        
        tabla = soup.select_one(selector_tabla)
        if not tabla:
            logger.error(f"No se encontró tabla con selector: {selector_tabla}")
            return []
        
        # Extraer headers
        headers = []
        header_row = tabla.select_one('thead tr')
        if header_row:
            headers = [limpiar_texto(th.get_text()) for th in header_row.select('th')]
        
        # Si no hay headers en thead, buscar en primera fila
        if not headers:
            primera_fila = tabla.select_one('tr')
            if primera_fila:
                headers = [limpiar_texto(th.get_text()) for th in primera_fila.select('th')]
        
        # Extraer filas
        filas = tabla.select('tbody tr') if tabla.select('tbody') else tabla.select('tr')[1:]
        
        datos = []
        for fila in filas:
            celdas = fila.select('td')
            if celdas and headers and len(celdas) == len(headers):
                fila_dict = {}
                for i, celda in enumerate(celdas):
                    fila_dict[headers[i]] = limpiar_texto(celda.get_text())
                datos.append(fila_dict)
        
        logger.info(f"Extraídas {len(datos)} filas de tabla")
        return datos

# =============================================================================
# EXPORTADORES DE DATOS
# =============================================================================

class ExportadorDatos:
    """Clase para exportar datos a diferentes formatos"""
    
    @staticmethod
    def a_json(datos: List[Dict], nombre_archivo: str) -> str:
        """
        Exporta datos a formato JSON.
        
        Args:
            datos: Lista de diccionarios con los datos
            nombre_archivo: Nombre del archivo (sin extensión)
            
        Returns:
            Ruta del archivo creado
        """
        crear_directorio_salida()
        ruta = os.path.join(DIRECTORIO_SALIDA, f"{nombre_archivo}.json")
        
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2, default=str)
        
        logger.info(f"Datos exportados a JSON: {ruta}")
        return ruta
    
    @staticmethod
    def a_csv(datos: List[Dict], nombre_archivo: str) -> str:
        """
        Exporta datos a formato CSV.
        
        Args:
            datos: Lista de diccionarios con los datos
            nombre_archivo: Nombre del archivo (sin extensión)
            
        Returns:
            Ruta del archivo creado
        """
        if not datos:
            logger.warning("No hay datos para exportar")
            return ""
        
        crear_directorio_salida()
        ruta = os.path.join(DIRECTORIO_SALIDA, f"{nombre_archivo}.csv")
        
        # Obtener todas las claves posibles
        claves = set()
        for item in datos:
            claves.update(item.keys())
        claves = sorted(list(claves))
        
        with open(ruta, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=claves)
            writer.writeheader()
            writer.writerows(datos)
        
        logger.info(f"Datos exportados a CSV: {ruta}")
        return ruta
    
    @staticmethod
    def a_excel(datos: List[Dict], nombre_archivo: str) -> Optional[str]:
        """
        Exporta datos a formato Excel (.xlsx).
        
        Args:
            datos: Lista de diccionarios con los datos
            nombre_archivo: Nombre del archivo (sin extensión)
            
        Returns:
            Ruta del archivo creado o None si openpyxl no está instalado
        """
        if not EXCEL_DISPONIBLE:
            logger.warning("openpyxl no está instalado. Instala con: pip install openpyxl")
            return None
        
        if not datos:
            logger.warning("No hay datos para exportar")
            return None
        
        crear_directorio_salida()
        ruta = os.path.join(DIRECTORIO_SALIDA, f"{nombre_archivo}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Obtener headers
        claves = sorted(list(datos[0].keys()))
        
        # Escribir headers
        for col, clave in enumerate(claves, 1):
            ws.cell(row=1, column=col, value=clave)
        
        # Escribir datos
        for row, item in enumerate(datos, 2):
            for col, clave in enumerate(claves, 1):
                valor = item.get(clave, '')
                ws.cell(row=row, column=col, value=str(valor) if valor else '')
        
        wb.save(ruta)
        logger.info(f"Datos exportados a Excel: {ruta}")
        return ruta

# =============================================================================
# EJEMPLO DE USO Y DEMO
# =============================================================================

async def demo_scraper_noticias():
    """
    Demostración del scraper de noticias asíncrono.
    """
    print("\n" + "="*60)
    print("📰 DEMO: SCRAPER DE NOTICIAS (ASYNC)")
    print("="*60)
    
    scraper = ScraperNoticias()
    try:
        urls_demo = [
            "https://elpais.com/tecnologia/",
            "https://www.bbc.com/news/technology"
        ]
        
        todas_noticias = []
        
        for url in urls_demo:
            print(f"\n🔍 Extrayendo noticias de: {url}")
            noticias = await scraper.extraer_noticias(url, max_noticias=5)
            
            for i, noticia in enumerate(noticias, 1):
                print(f"\n  [{i}] {noticia.get('titulo', 'Sin título')[:80]}...")
        
            todas_noticias.extend(noticias)
        
        if todas_noticias:
            ExportadorDatos.a_json(todas_noticias, "noticias_demo")
            print(f"\n✅ Exportadas {len(todas_noticias)} noticias")
            
        return todas_noticias
    finally:
        await scraper.cerrar()


async def demo_scraper_ecommerce():
    """
    Demostración del scraper de e-commerce paralelo.
    """
    print("\n" + "="*60)
    print("🛒 DEMO: SCRAPER DE E-COMMERCE (PARALLEL)")
    print("="*60)
    
    scraper = ScraperEcommerce()
    try:
        urls_productos = [
            # Amazon / MercadoLibre URLs de ejemplo
        ]
        
        if not urls_productos:
            print("\n⚠️  Para probar el scraper de e-commerce, agrega URLs en scraper.py")
            return []
        
        productos = await scraper.monitorear_productos(urls_productos)
        
        for producto in productos:
            print(f"\n📦 {producto.nombre[:60]}...")
            print(f"   💰 Precio: {producto.precio}")
        
        if productos:
            datos = [asdict(p) for p in productos]
            ExportadorDatos.a_json(datos, "productos_demo")
            print(f"\n✅ Exportados {len(productos)} productos")
            
        return productos
    finally:
        await scraper.cerrar()


def demo_scraper_generico():
    """
    Demostración del scraper genérico.
    Muestra cómo extraer datos con selectores personalizados.
    """
    print("\n" + "="*60)
    print("🔧 DEMO: SCRAPER GENÉRICO")
    print("="*60)
    
    scraper = ScraperGenerico()
    
    # Ejemplo: extraer datos de una página con selectores específicos
    # El usuario puede configurar sus propios selectores
    
    print("\n📌 El scraper genérico permite extraer datos de cualquier página")
    print("   usando selectores CSS personalizados.\n")
    
    print("   Ejemplo de uso:")
    print("""
    configuracion = {
        'titulo': 'h1',
        'precio': '.price-amount',
        'descripcion': '.product-description',
        'imagen': {'selector': '.main-image img', 'atributo': 'src'}
    }
    
    datos = scraper.extraer_datos('https://ejemplo.com/producto', configuracion)
    """)
    
    print("\n   Para extraer tablas:")
    print("""
    datos_tabla = scraper.extraer_tabla(
        'https://ejemplo.com/datos',
        'table.datos'
    )
    """)


# Nota: La ejecución principal asíncrona se realiza ahora desde main.py
